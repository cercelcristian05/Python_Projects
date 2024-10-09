from openpyxl import Workbook

wb = Workbook()
ws = wb.active

definition = ["Test length", "Exercise", "Total of Set", "Weight", "Total of Reps"]

x = 1; y = 1; previous_x=0; previous_y=0 # Position Worksheet
all_exercise = {}; all_sets = {}; all_weight={}; all_reps={}  # Dictionary to store sets for each exercise

def createTitle():
    global previous_x
    ws.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y+3) 
    ws.cell(row=x, column=y, value=definition[0])
    previous_x = x

def createSubTitle():
    global previous_x
    for i, title in enumerate(definition[1:], start=1):
        ws.cell(row=previous_x+1, column=y+i-1, value=title)
    previous_x += 1

def createExercise():
    global all_exercise
    all_exercise = []  # Reset the list for new inputs
    range_exercise = int(input("Input maximum number of exercises: "))
    for _ in range(range_exercise):
        count_exercise = input("Enter the name of exercise: ")
        all_exercise.append(count_exercise)

def createValueforExercise():
    global all_sets, all_weight, all_reps 
    all_sets = {}; all_weight= {}; all_reps = {}  # Reset the dictionaries for new inputs
    for exercise in all_exercise:
        sets = []; weights = []; reps = []
        range_set = int(input(f"Input maximum number of sets for {exercise}: "))
        for _ in range(range_set):
            count_set = input(f"What's your set for {exercise} ?: ")
            sets.append(count_set)
            count_weight = input(f"How much you lift for {exercise} you do?: ")
            weights.append(count_weight)
            count_rep = input(f"How many rep you do for {exercise} ?: ")
            reps.append(count_rep)
        all_sets[exercise] = sets
        all_weight[exercise] = weights
        all_reps[exercise] = reps       

def introduceValue():
    global previous_x
    for exercise in all_exercise:
        sets = all_sets[exercise]
        weights = all_weight[exercise]
        reps = all_reps[exercise]

        merge_start = previous_x + 1
        merge_end = merge_start + len(sets) - 1

        # Merge cells for exercise name
        ws.merge_cells(start_row=merge_start, start_column=y, end_row=merge_end, end_column=y)
        ws.cell(row=merge_start, column=y, value=exercise).alignment = Workbook(horizontal='center',vertical='center')

        # Write sets, weights, and reps in separate columns
        for j, (set_value, weight_value, rep_value) in enumerate(zip(sets, weights, reps)):
            ws.cell(row=merge_start + j, column=y + 1, value=set_value).alignment = Workbook(horizontal='center',vertical='center')
            ws.cell(row=merge_start + j, column=y + 2, value=weight_value).alignment = Workbook(horizontal='center',vertical='center')
            ws.cell(row=merge_start + j, column=y + 3, value=rep_value).alignment = Workbook(horizontal='center',vertical='center')

        previous_x = merge_end

def createTable():
    createTitle() 
    createSubTitle()
    createExercise() 
    createValueforExercise()
    introduceValue()

createTable()
# Save the workbook
wb.save('your_file.xlsx')
